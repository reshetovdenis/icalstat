import sys
import os
import argparse
from datetime import datetime, timedelta
import objc
from Foundation import NSObject, NSDate, NSRunLoop
from AppKit import NSColorSpace
from EventKit import (
    EKEventStore,
    EKEntityTypeEvent,
    EKAuthorizationStatusNotDetermined,
    EKAuthorizationStatusAuthorized,
    EKAuthorizationStatusDenied,
)
import threading
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.axis import ChartLines
from openpyxl.drawing.fill import ColorChoice


# Helper function to format dates as YYYY-MM-DD
def format_date(date_obj):
    return date_obj.strftime("%Y-%m-%d")


# Helper function to round a number to specified decimal places
def round_number(n, decimal_places):
    return round(n, decimal_places)


# Handler to find a record by date and calendar
def find_record(summary_data, search_date, search_calendar):
    for record in summary_data:
        if record['eventDate'] == search_date and record['calendarName'] == search_calendar:
            return record
    return None


class CalendarSummaryReport(NSObject):
    def init(self):
        self = objc.super(CalendarSummaryReport, self).init()
        if self is None:
            return None
        self.store = EKEventStore.alloc().init()
        self.summary_data = []
        self.detailed_data = []  # Added to track individual event data
        self.access_event = threading.Event()  # For synchronization
        self.access_granted = False
        self.calendar_colors = {}  # Dictionary to store calendar colors
        return self

    def request_access(self):
        authorization_status = EKEventStore.authorizationStatusForEntityType_(EKEntityTypeEvent)
        if authorization_status == EKAuthorizationStatusNotDetermined:
            self.store.requestAccessToEntityType_completion_(EKEntityTypeEvent, self.access_callback)
            timeout = 10  # seconds
            end_time = datetime.now() + timedelta(seconds=timeout)
            while not self.access_event.is_set() and datetime.now() < end_time:
                NSRunLoop.currentRunLoop().runMode_beforeDate_('default', NSDate.dateWithTimeIntervalSinceNow_(0.1))
            return self.access_granted
        elif authorization_status == EKAuthorizationStatusAuthorized:
            return True
        else:
            print("Access to Calendar is denied. Please enable it in System Preferences.")
            return False

    def access_callback(self, granted, error):
        self.access_granted = granted
        self.access_event.set()

    def generate_report(self, last_days=0):
        # Define the date range
        end_date = datetime.now()
        start_date = end_date - timedelta(days=last_days)

        # Set time to start of day and end of day
        start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=0)

        if last_days > 0:
            start_date = (end_date - timedelta(days=last_days)).replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = end_date - timedelta(days=1)
            end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=0)

        # Convert to NSDate
        start_nsdate = self.python_date_to_nsdate(start_date)
        end_nsdate = self.python_date_to_nsdate(end_date)

        # Get all calendars
        calendars = self.store.calendarsForEntityType_(EKEntityTypeEvent)

        if not calendars:
            print("No calendars found. Please ensure you have at least one calendar in the Calendar app.")
            sys.exit(1)

        # Extract and store calendar colors
        for cal in calendars:
            cal_name = cal.title()
            ns_color = cal.color()
            hex_color = self.nscolor_to_hex(ns_color)
            self.calendar_colors[cal_name] = hex_color

            predicate = self.store.predicateForEventsWithStartDate_endDate_calendars_(
                start_nsdate, end_nsdate, [cal]
            )
            events = self.store.eventsMatchingPredicate_(predicate)

            for event in events:
                evt_start = event.startDate()
                evt_end = event.endDate()
                duration_seconds = evt_end.timeIntervalSinceDate_(evt_start)
                duration_hours = duration_seconds / 3600
                evt_date = format_date(self.nsdate_to_python_date(evt_start))
                event_name = event.title() if event.title() else "Unnamed Event"

                self.detailed_data.append({
                    'eventDate': evt_date,
                    'calendarName': cal_name,
                    'eventName': event_name,
                    'totalDuration': duration_hours,
                })

                existing_record = find_record(self.summary_data, evt_date, cal_name)
                if existing_record:
                    existing_record['totalDuration'] += duration_hours
                else:
                    self.summary_data.append({
                        'eventDate': evt_date,
                        'calendarName': cal_name,
                        'totalDuration': duration_hours,
                    })

        self.summary_data.sort(key=lambda x: x['eventDate'])
        self.detailed_data.sort(key=lambda x: (x['calendarName'], x['eventDate']))

    def export_to_xlsx(self, days_in_period, output_path=None):
        headers = ['Date', 'Calendar', 'Total Duration Hours']
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        bold_font = Font(bold=True)
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header).font = bold_font

        for row_num, record in enumerate(self.summary_data, 2):
            ws.cell(row=row_num, column=1, value=record['eventDate'])
            ws.cell(row=row_num, column=2, value=record['calendarName'])
            ws.cell(row=row_num, column=3, value=record['totalDuration'])

            # Apply background color based on calendar
            calendar_name = record['calendarName']
            if calendar_name in self.calendar_colors:
                hex_color = self.calendar_colors[calendar_name]
                fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                for col in range(1, 4):  # Assuming 3 columns
                    ws.cell(row=row_num, column=col).fill = fill

        # Add bar chart sheet
        self.add_bar_chart_sheet(wb)

        # Add calendar-specific tabs
        self.add_calendar_tabs(wb, days_in_period)

        output_path = output_path or os.path.join(os.path.expanduser("~/Desktop"), "Calendar_Summary_Report.xlsx")
        wb.save(output_path)
        print(f"Report saved to {output_path}")

    def add_calendar_tabs(self, wb, days_in_period):
        events_by_calendar = {}
        for event in self.detailed_data:
            calendar_name = event['calendarName']
            if calendar_name not in events_by_calendar:
                events_by_calendar[calendar_name] = []
            events_by_calendar[calendar_name].append(event)

        for calendar_name, events in events_by_calendar.items():
            aggregated_data = {}
            for event in events:
                event_name = event['eventName']
                if event_name not in aggregated_data:
                    aggregated_data[event_name] = 0
                aggregated_data[event_name] += event['totalDuration']

            ws = wb.create_sheet(title=calendar_name[:31])
            headers = ["Event Name", "Total Duration Hours", "Hours per Day"]
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header).font = Font(bold=True)

            for row_num, (event_name, total_duration) in enumerate(aggregated_data.items(), 2):
                ws.cell(row=row_num, column=1, value=event_name)
                ws.cell(row=row_num, column=2, value=round(total_duration, 2))
                ws.cell(row=row_num, column=3, value=round(total_duration / days_in_period, 2))

            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                    adjusted_width = max_length + 2
                    ws.column_dimensions[column_letter].width = adjusted_width

    def add_bar_chart_sheet(self, wb):
        chart_ws = wb.create_sheet(title="By day")
        chart_headers = ['Date'] + list(set(record['calendarName'] for record in self.summary_data))
        chart_ws.append(chart_headers)

        chart_data = {}
        for record in self.summary_data:
            date = record['eventDate']
            calendar_name = record['calendarName']
            duration = record['totalDuration']
            if date not in chart_data:
                chart_data[date] = {cal: 0 for cal in chart_headers[1:]}
            chart_data[date][calendar_name] += duration

        for date, data in sorted(chart_data.items()):
            row = [date] + [data.get(cal, 0) for cal in chart_headers[1:]]
            chart_ws.append(row)

        chart = BarChart()
        chart.type = "col"
        chart.style = 12
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.legend.position = 't'

        data = Reference(chart_ws, min_col=2, max_col=len(chart_headers), min_row=1, max_row=len(chart_data) + 1)
        categories = Reference(chart_ws, min_col=1, min_row=2, max_row=len(chart_data) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.shape = 4
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.y_axis.majorGridlines = ChartLines(spPr=GraphicalProperties(ln=LineProperties(prstDash='dot')))

        # Apply colors to the series based on the calendar hex colors
        for i, series in enumerate(chart.series, start=1):
            if i - 1 < len(chart_headers) - 1:
                calendar_name = chart_headers[i]
                if calendar_name in self.calendar_colors:
                    hex_color = self.calendar_colors[calendar_name]
                    # Convert hex color to RGB for openpyxl
                    r = hex_color[:2]
                    g = hex_color[2:4]
                    b = hex_color[4:]
                    rgb_color = f"{r}{g}{b}"
                    series.graphicalProperties.solidFill = rgb_color

        # Add the chart to the chart sheet
        chart_ws.add_chart(chart, "E5")

    def python_date_to_nsdate(self, py_date):
        epoch = datetime(2001, 1, 1)
        delta = py_date - epoch
        return NSDate.dateWithTimeIntervalSinceReferenceDate_(delta.total_seconds())

    def nsdate_to_python_date(self, ns_date):
        epoch = datetime(2001, 1, 1)
        interval = ns_date.timeIntervalSinceReferenceDate()
        return epoch + timedelta(seconds=interval)

    def nscolor_to_hex(self, ns_color):
        try:
            r = int(ns_color.redComponent() * 255)
            g = int(ns_color.greenComponent() * 255)
            b = int(ns_color.blueComponent() * 255)
            return f"{r:02X}{g:02X}{b:02X}"
        except Exception:
            return "D3D3D3"


def parse_arguments():
    parser = argparse.ArgumentParser(
        description="Generate a summary report of Apple Calendar events."
    )
    parser.add_argument(
        'days',
        type=int,
        nargs='?',
        default=1,
        help='Number of days to include in the report (default: 1, which means today).'
    )
    parser.add_argument(
        '-o', '--output',
        type=str,
        default=None,
        help='Path to save the XLSX report (default: Desktop/Calendar_Summary_Report.xlsx).'
    )
    return parser.parse_args()


def main():
    args = parse_arguments()
    last_days = args.days
    output_path = args.output

    days_in_period = max(last_days, 1)

    report = CalendarSummaryReport.alloc().init()

    if not report.request_access():
        sys.exit(1)

    report.generate_report(last_days=last_days)
    report.export_to_xlsx(days_in_period=days_in_period, output_path=output_path)


if __name__ == "__main__":
    main()